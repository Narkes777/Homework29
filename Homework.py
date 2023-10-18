import openpyxl
import requests

# # Task 1
# book = openpyxl.Workbook()

# sheet = book.active

# sheet['A1'] = 'Имя'
# sheet['B1'] = 'Email'

# users_data = [
#     ['User1', 'user1@mail.ru'],
#     ['User2', 'user2@mail.ru'],
#     ['User3', 'user3@mail.ru']
# ]

# for row, user in enumerate(users_data, start=2):
#     sheet[f'A{row}'] = user[0]
#     sheet[f'B{row}'] = user[1]

# book.save('users.xlsx')

# print("Данные успешно записаны в файл users.xlsx")

# # Task 2
# book = openpyxl.load_workbook('users.xlsx')

# sheet = book.active

# for row in sheet.iter_rows(min_row=2, values_only=True):
#     name, email = row
#     print(f"Имя: {name}, Email: {email}")

# book.close()

# # JSON
# # Task 1
# url = 'https://jsonplaceholder.typicode.com/users'

# response = requests.get(url)

# if response.status_code == 200:
#     users = response.json()

#     for user in users:
#         print(f"Имя: {user['name']}, Почта: {user['email']}")
# else:
#     print(f"Ошибка при выполнении GET-запроса: {response.status_code}")

# # Task 2
# url = 'https://jsonplaceholder.typicode.com/users'

# response = requests.get(url)

# if response.status_code == 200:
#     users = response.json()

#     for user in users:
#         print(user['name'])
# else:
#     print(f"Ошибка при выполнении GET-запроса: {response.status_code}")
